"""Instrumento de κ humano (Fases 14/18).

O que estes testes protegem:

1. **O anotador não pode oferecer vocabulário que não existe.** As listas do menu são explícitas
   (parsear texto livre seria frágil), e o preço disso é risco de DRIFT: se o catálogo mudar e as
   listas não, o humano anotaria com tokens inválidos e o κ mediria ruído. Este é o teste que
   justifica a escolha de hardcodar.
2. **Nada é preenchido por máquina.** É a fronteira do projeto: gerar specs e chamá-las de humanas
   seria fabricar evidência. O `kappa` se recusa a rodar sem input humano, e há teste para isso.
3. **Nenhum default nas escolhas.** Uma opção pré-selecionada seria viés silencioso empurrando o
   anotador para uma resposta — o oposto do que uma anotação cega deve fazer.
"""
from __future__ import annotations

import importlib.util
import json

import pytest

import anotar_humano as ah
from rodoquery.gold import Spec
from rodoquery.sistema_antt import CATALOGO_ANTT

# openpyxl e' extra (`pip install -e .[xlsx]`): so o caminho da planilha depende dele, e pular
# esses testes nao pode arrastar os do anotador guiado junto.
precisa_xlsx = pytest.mark.skipif(importlib.util.find_spec("openpyxl") is None,
                                  reason="openpyxl não instalado (extra `xlsx`)")


# ------------------------------------------------------------------ drift contra o catalogo
@pytest.mark.parametrize("metrica", ah.METRICAS)
def test_metrica_do_menu_existe_no_catalogo(metrica):
    assert metrica in CATALOGO_ANTT


@pytest.mark.parametrize("token", ah.GROUP_BY)
def test_token_de_group_by_existe_no_catalogo(token):
    assert token in CATALOGO_ANTT


@pytest.mark.parametrize("dim", sorted(ah.VALORES))
def test_dimensao_filtravel_existe_no_catalogo(dim):
    assert dim in CATALOGO_ANTT


def test_valores_categoricos_existem_no_catalogo():
    for dim, vals in ah.VALORES.items():
        if dim == "plaza__categoria_eixo":
            continue          # o catálogo descreve a faixa ('2' a '20'), não enumera
        for v in vals:
            assert v in CATALOGO_ANTT, f"{dim}={v} não está no catálogo"


def test_menu_nao_oferece_metrica_alem_das_tres():
    """O catálogo da ANTT tem 3 métricas de usuário — nem 2 nem 4."""
    assert len(ah.METRICAS) == 3


# --------------------------------------------------------------------------- escolha numerada
def _com_entradas(monkeypatch, *respostas):
    it = iter(respostas)
    monkeypatch.setattr("builtins.input", lambda *a: next(it))


def test_menu_escolha_unica(monkeypatch):
    _com_entradas(monkeypatch, "2")
    assert ah._menu("t", ["a", "b", "c"], multi=False, permite_vazio=False) == ["b"]


def test_menu_multipla(monkeypatch):
    _com_entradas(monkeypatch, "1,3")
    assert ah._menu("t", ["a", "b", "c"], multi=True, permite_vazio=False) == ["a", "c"]


def test_menu_vazio_quando_permitido(monkeypatch):
    _com_entradas(monkeypatch, "")
    assert ah._menu("t", ["a"], multi=True, permite_vazio=True) == []


def test_menu_reprova_entrada_invalida_ate_acertar(monkeypatch):
    """Não aceita lixo silenciosamente: uma spec inválida contaminaria o κ."""
    _com_entradas(monkeypatch, "9", "abc", "", "1")     # fora do range, não-número, vazio, ok
    assert ah._menu("t", ["a", "b"], multi=False, permite_vazio=False) == ["a"]


def test_menu_reprova_multipla_quando_e_escolha_unica(monkeypatch):
    _com_entradas(monkeypatch, "1,2", "2")
    assert ah._menu("t", ["a", "b"], multi=False, permite_vazio=False) == ["b"]


# ------------------------------------------------------------------------------ filtro where
def test_where_gera_sintaxe_do_metricflow(monkeypatch):
    dims = sorted(ah.VALORES)
    i = dims.index("plaza__sentido") + 1
    _com_entradas(monkeypatch, str(i), "1")             # dimensão, 1º valor (Crescente)
    assert ah._montar_where() == "{{ Dimension('plaza__sentido') }} = 'Crescente'"


def test_where_pode_ser_nenhum(monkeypatch):
    _com_entradas(monkeypatch, "")
    assert ah._montar_where() is None


def test_where_gerado_bate_com_a_sintaxe_documentada_no_catalogo():
    """O catálogo mostra o formato; se o anotador gerar outro, o gold não compila."""
    assert "{{ Dimension('plaza__tipo_cobranca') }} = 'Automática'" in CATALOGO_ANTT


# ------------------------------------------------- a fronteira: nada e preenchido por maquina
def test_kappa_se_recusa_sem_input_humano(tmp_path, monkeypatch):
    folha = tmp_path / "folha.jsonl"
    folha.write_text('{"id": "x", "estrato": "e", "pergunta_nl": "p", "spec_humano": null}\n',
                     encoding="utf-8")
    monkeypatch.setattr(ah, "FOLHA", folha)
    with pytest.raises(SystemExit, match="se recusa a inventar"):
        ah.kappa()


def test_kappa_aborta_com_item_fora_do_golden_selado(tmp_path, monkeypatch):
    """Anotar contra referência que as Fases 14/15 descartaram mediria defeito conhecido."""
    folha = tmp_path / "folha.jsonl"
    spec = {"metrics": ["traffic_volume"], "group_by": [], "where": None,
            "order_by": [], "limit": None, "ordenado": False}
    folha.write_text(json.dumps({"id": "id_que_nao_existe", "estrato": "e",
                                 "pergunta_nl": "p", "spec_humano": spec}) + "\n",
                     encoding="utf-8")
    monkeypatch.setattr(ah, "FOLHA", folha)
    with pytest.raises(SystemExit, match="NÃO estão no golden selado"):
        ah.kappa()


def test_anotar_nao_mostra_o_estrato_do_autor(tmp_path, monkeypatch, capsys):
    """O nome do estrato é metade da resposta — imprimi-lo infla o κ.

    `abstencao` diz "não responda"; `ranking` diz "é um ranking"; `grao_temporal` diz "agrupe por
    tempo". O anotador é cego à spec do autor, mas via o rótulo do autor logo acima da pergunta.
    """
    folha = tmp_path / "folha.jsonl"
    folha.write_text(json.dumps({"id": "x1", "estrato": "abstencao", "pergunta_nl": "P?",
                                 "spec_humano": None}, ensure_ascii=False) + "\n",
                     encoding="utf-8")
    monkeypatch.setattr(ah, "FOLHA", folha)
    _com_entradas(monkeypatch, "", "4", "s")            # ENTER, ABSTENHO, confirma
    ah.anotar()
    assert "abstencao" not in capsys.readouterr().out


def test_anotar_percorre_em_ordem_embaralhada(tmp_path, monkeypatch):
    """A folha é escrita agrupada por estrato: a ORDEM sozinha entrega o rótulo.

    Cinco abstenções seguidas e o 5º item se denuncia mesmo com o nome do estrato escondido.
    """
    folha = tmp_path / "folha.jsonl"
    folha.write_text("".join(
        json.dumps({"id": f"x{i}", "estrato": "abstencao", "pergunta_nl": f"P{i}?",
                    "spec_humano": None}, ensure_ascii=False) + "\n" for i in range(20)),
        encoding="utf-8")
    monkeypatch.setattr(ah, "FOLHA", folha)
    vistos = []
    respostas = iter(["q"])
    monkeypatch.setattr("builtins.input", lambda *a: next(respostas))
    monkeypatch.setattr("builtins.print", lambda *a, **k: vistos.append(" ".join(map(str, a))))
    ah.anotar()
    primeira = next(v for v in vistos if "P" in v and "?" in v)
    assert "P0?" not in primeira, "percorreu na ordem do arquivo — a ordem vaza o estrato"


def test_kappa_registra_humano_respondendo_item_de_abstencao(tmp_path, monkeypatch):
    """A discordância mais informativa não pode DERRUBAR o κ.

    O humano responder um item que o autor marcou fora-de-escopo é exatamente o sinal que a
    amostra existe para captar (5 dos 40 itens são do estrato `abstencao`). Herdar o estrato do
    autor no lado B fazia o `ItemGolden` recusar essa spec ("abstencao ⟹ metrics vazio") e o
    `kappa` estourava DEPOIS da hora de anotação — com o convite implícito de "consertar"
    editando a label do humano, que é o único conserto que invalidaria a medição.
    """
    alvo = next(it for it in ah._golden_selado() if it.estrato == "abstencao")
    spec = {"metrics": ["traffic_volume"], "group_by": [], "where": None,
            "order_by": [], "limit": None, "ordenado": False}
    folha = tmp_path / "folha.jsonl"
    folha.write_text(json.dumps({"id": alvo.id, "estrato": alvo.estrato,
                                 "pergunta_nl": alvo.pergunta_nl, "spec_humano": spec},
                                ensure_ascii=False) + "\n", encoding="utf-8")
    monkeypatch.setattr(ah, "FOLHA", folha)
    monkeypatch.setattr(ah, "D", tmp_path)

    ah.kappa()                                   # antes: ValueError e nenhum artefato

    saida = json.loads((tmp_path / "kappa_humano.json").read_text(encoding="utf-8"))
    assert saida["n_anotados_por_humano"] == 1
    # e a discordância foi REGISTRADA, não engolida
    assert saida["concordancia_spec"]["discordantes"] == [alvo.id]
    assert saida["concordancia_spec"]["concordancia_metrica"] == 0.0


def test_folha_atual_esta_anotada_e_sem_orfaos():
    """A guarda viva, virada.

    Ate 28/07/2026 este teste exigia a folha VAZIA e o proprio autor escreveu no lugar: "se um
    humano anotar de verdade, este assert cai — e ai e para trocar por um teste do kappa". Um
    humano anotou os 40. O invariante agora e o oposto: a folha esta completa, sem orfaos, e as
    specs sao carregaveis como `Spec` — se alguem a esvaziar ou corromper, isto cai.
    """
    if not ah.FOLHA.exists():
        pytest.skip("folha ainda não gerada neste checkout")
    linhas = [json.loads(x) for x in ah.FOLHA.read_text(encoding="utf-8").splitlines()
              if x.strip() and not x.startswith("#")]
    selado = {it.id for it in ah._golden_selado()}
    assert [d["id"] for d in linhas if d["id"] not in selado] == []
    faltando = [d["id"] for d in linhas if not d.get("spec_humano")]
    assert not faltando, f"{len(faltando)} itens perderam a anotação humana: {faltando[:5]}"
    for d in linhas:
        Spec(**d["spec_humano"])          # spec malformada contaminaria o κ silenciosamente


def test_kappa_humano_publicado_bate_com_a_folha():
    """O artefato publicado nao pode divergir da folha que o gerou.

    `reports/fase14/kappa_humano.json` e a evidencia que o README cita. Se a folha mudar e o
    artefato nao for regerado, o projeto passa a publicar um numero que nao corresponde mais ao
    dado — o tipo de deriva que so aparece quando alguem confere a mao.
    """
    art = ah.D / "kappa_humano.json"
    if not art.exists():
        pytest.skip("κ humano ainda não calculado neste checkout")
    saida = json.loads(art.read_text(encoding="utf-8"))
    linhas = [json.loads(x) for x in ah.FOLHA.read_text(encoding="utf-8").splitlines()
              if x.strip() and not x.startswith("#")]
    anotados = sum(1 for d in linhas if d.get("spec_humano"))
    assert saida["n_anotados_por_humano"] == anotados
    assert saida["concordancia_spec"]["n_pares"] == anotados


# --------------------------------------------------------------------------- planilha (XLSX)
def _folha_de_teste(tmp_path, monkeypatch, n=6):
    """Folha temporaria com ids REAIS do golden selado (o import mapeia por hash do id)."""
    itens = ah._golden_selado()[:n]
    folha = tmp_path / "folha.jsonl"
    folha.write_text("".join(
        json.dumps({"id": it.id, "estrato": it.estrato, "pergunta_nl": it.pergunta_nl,
                    "spec_humano": None}, ensure_ascii=False) + "\n" for it in itens),
        encoding="utf-8")
    monkeypatch.setattr(ah, "FOLHA", folha)
    return itens


@precisa_xlsx
def test_xlsx_nao_vaza_o_estrato_em_lugar_nenhum(tmp_path, monkeypatch):
    """Nem em coluna, nem no `id`.

    Os ids do golden sao PREFIXADOS pelo estrato (`abstencao_antt_23`), entao uma coluna de id
    entregaria a resposta tao bem quanto um rotulo. Por isso a planilha carrega um hash opaco.
    """
    from openpyxl import load_workbook
    itens = _folha_de_teste(tmp_path, monkeypatch)
    xl = tmp_path / "p.xlsx"
    ah.exportar_xlsx(xl)
    wb = load_workbook(xl)

    # (a) nenhum id do golden, em aba nenhuma — o id sozinho ja' e' o estrato
    todo = " ".join(str(c.value) for aba in wb.sheetnames for lin in wb[aba].iter_rows()
                    for c in lin if c.value is not None)
    for it in itens:
        assert it.id not in todo, f"a planilha vaza o id '{it.id}' (prefixado pelo estrato)"

    # (b) nenhum NOME de estrato na aba de anotação. A checagem e' aqui e nao no arquivo todo
    # de proposito: a aba `Instrucoes` diz "'por mes' e 'por praca' NAO sao ranking", que e'
    # REGRA DA TAREFA — vale para todos os itens e nao identifica nenhum. Vazamento e' o rotulo
    # colado a UM item; regra geral e' o que o anotador tem direito de saber.
    ws = wb["Anotacao"]
    celulas = " ".join(str(c.value) for lin in ws.iter_rows() for c in lin if c.value is not None)
    for estrato in ("abstencao", "ranking", "grao_temporal", "valor_categorico", "join_grao",
                    "coalesce_nulo", "metrica_derivada", "controle_trivial"):
        assert estrato not in celulas, f"a aba de anotação vaza o estrato '{estrato}'"


@precisa_xlsx
def test_xlsx_sai_sem_nenhuma_resposta_preenchida(tmp_path, monkeypatch):
    """Uma celula pre-preenchida seria viés silencioso — a mesma regra dos menus do terminal."""
    from openpyxl import load_workbook
    itens = _folha_de_teste(tmp_path, monkeypatch)
    xl = tmp_path / "p.xlsx"
    ah.exportar_xlsx(xl)
    ws = load_workbook(xl)["Anotacao"]
    assert ws.max_row == len(itens) + 1
    for r in range(2, ws.max_row + 1):
        for c in range(4, len(ah.COLUNAS) + 1):        # da METRICA_1 em diante
            assert ws.cell(row=r, column=c).value is None


@precisa_xlsx
def test_xlsx_ida_e_volta_preserva_a_spec(tmp_path, monkeypatch):
    from openpyxl import load_workbook
    _folha_de_teste(tmp_path, monkeypatch)
    xl = tmp_path / "p.xlsx"
    ah.exportar_xlsx(xl)
    wb = load_workbook(xl)
    ws = wb["Anotacao"]
    ws.cell(row=2, column=4, value="ABSTENHO")
    ws.cell(row=3, column=4, value="automation_rate")
    ws.cell(row=3, column=6, value="plaza__praca")
    ws.cell(row=3, column=9, value="plaza__categoria_eixo")
    ws.cell(row=3, column=10, value="4")
    ws.cell(row=3, column=11, value="sim")
    ws.cell(row=3, column=12, value="automation_rate")
    ws.cell(row=3, column=13, value="sim")
    ws.cell(row=3, column=14, value=3)
    wb.save(xl)
    ah.importar_xlsx(xl)

    feitas = [d["spec_humano"] for d in ah._linhas_da_folha() if d.get("spec_humano")]
    assert {"metrics": [], "group_by": [], "where": None,
            "order_by": [], "limit": None, "ordenado": False} in feitas
    assert {"metrics": ["automation_rate"], "group_by": ["plaza__praca"],
            "where": "{{ Dimension('plaza__categoria_eixo') }} = '4'",
            "order_by": ["-automation_rate"], "limit": 3, "ordenado": True} in feitas
    assert len(feitas) == 2                       # o resto continua pendente, nao virou spec vazia


@precisa_xlsx
def test_importar_rejeita_token_fora_do_catalogo_sem_gravar_nada(tmp_path, monkeypatch):
    """A validação do Excel e' conselho, nao garantia: da' para colar por cima dela. Um token
    invalido que passasse viraria ruido DENTRO do κ, indistinguivel de discordancia real."""
    from openpyxl import load_workbook
    _folha_de_teste(tmp_path, monkeypatch)
    xl = tmp_path / "p.xlsx"
    ah.exportar_xlsx(xl)
    wb = load_workbook(xl)
    ws = wb["Anotacao"]
    ws.cell(row=2, column=4, value="traffic_volume")
    ws.cell(row=2, column=6, value="plaza__municipio")        # nao existe no catalogo
    ws.cell(row=3, column=4, value="traffic_volume")          # linha valida, na mesma planilha
    wb.save(xl)
    with pytest.raises(SystemExit, match="fora do catálogo"):
        ah.importar_xlsx(xl)
    # rejeicao e' TOTAL: nem a linha valida entra, para nao deixar import pela metade
    assert all(d["spec_humano"] is None for d in ah._linhas_da_folha())


@precisa_xlsx
def test_importar_rejeita_filtro_pela_metade(tmp_path, monkeypatch):
    from openpyxl import load_workbook
    _folha_de_teste(tmp_path, monkeypatch)
    xl = tmp_path / "p.xlsx"
    ah.exportar_xlsx(xl)
    wb = load_workbook(xl)
    ws = wb["Anotacao"]
    ws.cell(row=2, column=4, value="traffic_volume")
    ws.cell(row=2, column=9, value="plaza__sentido")          # dimensao sem valor
    wb.save(xl)
    with pytest.raises(SystemExit, match="têm de vir juntos"):
        ah.importar_xlsx(xl)


@precisa_xlsx
def test_xlsx_percorre_na_mesma_ordem_embaralhada_do_terminal(tmp_path, monkeypatch):
    """Os dois caminhos tem de ver a mesma sequencia: se divergirem, dois anotadores 'cegos' em
    meios diferentes teriam exposicoes diferentes, e o κ deixaria de ser comparavel."""
    from openpyxl import load_workbook
    _folha_de_teste(tmp_path, monkeypatch, n=12)
    xl = tmp_path / "p.xlsx"
    ah.exportar_xlsx(xl)
    ws = load_workbook(xl)["Anotacao"]
    na_planilha = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
    linhas = ah._linhas_da_folha()
    esperado = [linhas[i]["pergunta_nl"] for i in ah._ordem_embaralhada(linhas)]
    assert na_planilha == esperado


def test_amostra_e_estratificada_e_deterministica(tmp_path, monkeypatch):
    monkeypatch.setattr(ah, "FOLHA", tmp_path / "f1.jsonl")
    ah.amostra(40)
    a = (tmp_path / "f1.jsonl").read_text(encoding="utf-8")
    monkeypatch.setattr(ah, "FOLHA", tmp_path / "f2.jsonl")
    ah.amostra(40)
    assert a == (tmp_path / "f2.jsonl").read_text(encoding="utf-8")   # seed fixa
