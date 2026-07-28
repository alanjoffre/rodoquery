"""Fase 14 (#1) — INSTRUMENTO de anotação humana para o κ humano do golden RodoQuery.

**Por que isto e não uma anotação pronta.** O κ humano é dívida declarada desde a Fase 2. Eu NÃO
posso produzi-lo: gerar specs e chamá-las de humanas seria fabricar evidência — a fronteira que
este projeto não cruza. O que EU posso entregar é o instrumento que reduz o κ humano de "trabalho
de construção" a "~1h de um humano real clicando". É isso aqui.

O anotador humano vê APENAS a pergunta + o catálogo (cego às specs do autor-modelo), mapeia cada
uma, e este script:
  1. amostra estratificada de N itens do golden (determinística, para ser reproduzível);
  2. apresenta pergunta + catálogo, coleta a spec do humano num formato simples;
  3. ao final, calcula o κ humano × autor-modelo com o MESMO `concordancia_mapeamento` usado no
     κ de máquina — então o número é comparável, não um cálculo ad-hoc.

Status honesto enquanto ninguém rodou: `reports/fase14/kappa_humano.json` NÃO existe, e o README/
docs dizem "instrumento pronto, aguardando anotador humano". Nada é preenchido por máquina aqui.

Uso:
  python anotar_humano.py amostra [N]     # gera a folha de anotação (default 40 itens)
  python anotar_humano.py anotar          # anotador GUIADO no terminal (só digitar números)
  python anotar_humano.py xlsx [destino]  # exporta PLANILHA para anotar fora do terminal
  python anotar_humano.py importar <xlsx> # lê a planilha preenchida de volta para a folha
  python anotar_humano.py kappa           # calcula κ depois que o humano preencheu a folha

**O que o `anotar` faz e o que ele NÃO faz.** Ele monta o JSON a partir de escolhas numeradas,
salva a cada item (é retomável) e nunca pré-seleciona nada — não há opção "sugerida", porque um
default seria um viés silencioso empurrando o anotador. Ele **não** vê nem mostra a spec do
autor-modelo: a cegueira é o que faz o κ medir concordância independente em vez de quanto o
humano concorda com algo que acabou de ler.
"""
import hashlib
import json
import random
import sys
from pathlib import Path

from rodoquery.estat import cohen_kappa
from rodoquery.gold import Spec
from rodoquery.golden import (
    ESTRATO_ABSTENCAO,
    ESTRATOS_RESPONDIVEIS,
    ItemGolden,
    carregar,
    concordancia_mapeamento,
)
from rodoquery.proveniencia import carimbar
from rodoquery.sistema_antt import CATALOGO_ANTT

REPO = Path(__file__).resolve().parent
G = REPO / "golden"
D = REPO / "reports" / "fase14"
FOLHA = G / "anotacao_humana_antt.jsonl"          # o humano PREENCHE o campo "spec_humano"
D.mkdir(parents=True, exist_ok=True)

MODELO_SPEC = ('{"metrics": [...], "group_by": [...], "where": <str|null>, '
               '"order_by": [...], "limit": <int|null>, "ordenado": <bool>}  '
               '# metrics:[] = ABSTENHO')


def _golden_selado() -> list[ItemGolden]:
    """O golden SELADO ATUAL (test + dev), não o pool bruto `golden_antt.jsonl`.

    Por quê: o pool ainda contém itens que as Fases 14 e 15 **removeram** — 10 rankings com
    empate na zona de corte (gold não-determinístico) e 7 labels defeituosas achadas por
    auditoria adversarial. Amostrar de lá faria o anotador humano gastar tempo em itens cuja
    referência **já sabemos estar errada**, e o κ mediria defeito conhecido em vez de
    discordância genuína. A 1ª folha (25/07) tinha 3 desses em 40.
    """
    itens, vistos = [], set()
    for nome in ("golden_test_antt.jsonl", "golden_dev_antt.jsonl"):
        for it in carregar(G / nome):
            if it.id not in vistos:
                vistos.add(it.id)
                itens.append(it)
    return itens


def amostra(n: int) -> None:
    itens = _golden_selado()
    porestrato: dict[str, list] = {}
    for it in itens:
        porestrato.setdefault(it.estrato, []).append(it)
    rng = random.Random(42)
    escolha = []
    por = max(1, n // len(porestrato))
    for _, grupo in sorted(porestrato.items()):
        rng.shuffle(grupo)
        escolha += grupo[:por]
    linhas = []
    for it in escolha:
        linhas.append(json.dumps({
            "id": it.id, "estrato": it.estrato, "pergunta_nl": it.pergunta_nl,
            "spec_humano": None,          # <<< O HUMANO PREENCHE ISTO
        }, ensure_ascii=False))
    cab = (f"# FOLHA DE ANOTACAO HUMANA — {len(escolha)} itens\n"
           f"# Preencha 'spec_humano' em cada linha, cego as specs do autor. Formato:\n"
           f"#   {MODELO_SPEC}\n"
           f"# CATALOGO (o unico contexto permitido):\n"
           + "".join(f"#   {ln}\n" for ln in CATALOGO_ANTT.splitlines()))
    FOLHA.write_text(cab + "\n".join(linhas) + "\n", encoding="utf-8")
    print(f"folha gerada: {len(escolha)} itens estratificados -> {FOLHA}")
    print("Preencha 'spec_humano' em cada linha (cego), depois rode: python anotar_humano.py kappa")


# ---------------------------------------------------------------------------------------------
# Anotador guiado
#
# O vocabulário fica explícito aqui (e não parseado do catálogo) porque parsear texto livre é
# frágil. O preço disso é risco de DRIFT: se o catálogo mudar e estas listas não, o anotador
# ofereceria tokens que não existem mais. `tests/test_anotar_humano.py` trava isso — cada item
# abaixo tem de aparecer literalmente em CATALOGO_ANTT.
# ---------------------------------------------------------------------------------------------
METRICAS = ["traffic_volume", "automation_rate", "commercial_share"]
GROUP_BY = ["metric_time__day", "metric_time__week", "metric_time__month",
            "plaza__praca", "plaza__concessionaria", "plaza__sentido",
            "plaza__tipo_cobranca", "plaza__categoria_eixo", "plaza__tipo_de_veiculo"]
VALORES = {
    "plaza__tipo_cobranca": ["Automática", "Manual", "OCR/PLACA"],
    "plaza__tipo_de_veiculo": ["Comercial", "Passeio", "Moto"],
    "plaza__sentido": ["Crescente", "Decrescente"],
    "plaza__categoria_eixo": [str(i) for i in range(2, 21)],
}


def _menu(titulo: str, opcoes: list[str], multi: bool, permite_vazio: bool) -> list[str]:
    """Escolha numerada. NUNCA há opção pré-selecionada — default seria viés silencioso."""
    print(f"\n  {titulo}")
    for i, o in enumerate(opcoes, 1):
        print(f"    {i}. {o}")
    dica = "números separados por vírgula" if multi else "um número"
    dica += ", ENTER para nenhum" if permite_vazio else ""
    while True:
        cru = input(f"  > ({dica}): ").strip()
        if not cru:
            if permite_vazio:
                return []
            print("    ! obrigatório")
            continue
        try:
            idx = [int(x) for x in cru.replace(" ", "").split(",") if x]
            if not multi and len(idx) != 1:
                print("    ! escolha só um")
                continue
            if any(not 1 <= i <= len(opcoes) for i in idx):
                print(f"    ! fora do intervalo 1..{len(opcoes)}")
                continue
            return [opcoes[i - 1] for i in idx]
        except ValueError:
            print("    ! digite números")


def _sim(pergunta: str) -> bool:
    while True:
        r = input(f"  {pergunta} (s/n): ").strip().lower()
        if r in ("s", "sim"):
            return True
        if r in ("n", "nao", "não"):
            return False


def _montar_where() -> str | None:
    """Monta a sintaxe do MetricFlow, que é chata de digitar à mão e fácil de errar."""
    dims = sorted(VALORES)
    escolha = _menu("FILTRO (where) — dimensão:", [*dims, "(nenhum filtro)"], False, True)
    if not escolha or escolha[0] == "(nenhum filtro)":
        return None
    d = escolha[0]
    v = _menu(f"valor de {d}:", VALORES[d], False, False)[0]
    return f"{{{{ Dimension('{d}') }}}} = '{v}'"


def anotar() -> None:
    """Coleta a spec do HUMANO item a item. Salva a cada resposta — retomável."""
    if not FOLHA.exists():
        raise SystemExit("folha ausente — rode `python anotar_humano.py amostra` primeiro")
    texto = FOLHA.read_text(encoding="utf-8")
    cab = [x for x in texto.splitlines() if x.startswith("#")]
    linhas = [json.loads(x) for x in texto.splitlines() if x.strip() and not x.startswith("#")]

    def salvar() -> None:
        FOLHA.write_text("\n".join(cab) + "\n"
                         + "\n".join(json.dumps(d, ensure_ascii=False) for d in linhas) + "\n",
                         encoding="utf-8")

    # Ordem EMBARALHADA (semente fixa → retomável e auditável). A folha é escrita agrupada por
    # estrato, e a ordem sozinha entrega o rótulo do autor: depois de abster 4 vezes seguidas, o
    # 5º item se denuncia. Cegueira que o vizinho de linha desfaz não é cegueira.
    ordem = list(range(len(linhas)))
    random.Random(1337).shuffle(ordem)
    pendentes = [i for i in ordem if not linhas[i].get("spec_humano")]
    if not pendentes:
        print("Todos os itens já estão anotados. Rode: python anotar_humano.py kappa")
        return

    print(f"\n{'=' * 78}\nANOTACAO HUMANA — {len(pendentes)} de {len(linhas)} pendentes")
    print("Você vê apenas a pergunta e o catálogo. A spec do autor-modelo NÃO é mostrada:")
    print("é a cegueira que faz o κ medir concordância independente.")
    print("A qualquer momento:  c = ver catálogo   p = pular item   q = sair (salva)")
    print("=" * 78)

    for n, i in enumerate(pendentes, 1):
        d = linhas[i]
        # O estrato NÃO é impresso: `abstencao` diz "não responda", `ranking` diz "é um ranking",
        # `grao_temporal` diz "agrupe por tempo". É a classificação do autor — metade da resposta.
        # Mostrá-la infla o κ, que passa a medir concordância com uma dica.
        print(f"\n{'-' * 78}\n[{n}/{len(pendentes)}]")
        print(f"\n  PERGUNTA: {d['pergunta_nl']}\n")
        acao = input("  ENTER p/ anotar | c=catálogo | p=pular | q=sair: ").strip().lower()
        while acao == "c":
            print("\n" + CATALOGO_ANTT + "\n")
            acao = input("  ENTER p/ anotar | p=pular | q=sair: ").strip().lower()
        if acao == "q":
            salvar()
            print(f"\nSalvo. {sum(1 for x in linhas if x.get('spec_humano'))}/{len(linhas)} "
                  "anotados. Rode de novo para continuar.")
            return
        if acao == "p":
            continue

        met = _menu("MÉTRICA(S) — ou 'ABSTENHO' se nada no catálogo responde:",
                    [*METRICAS, "ABSTENHO (nenhuma métrica responde)"], True, False)
        if any(m.startswith("ABSTENHO") for m in met):
            spec = {"metrics": [], "group_by": [], "where": None,
                    "order_by": [], "limit": None, "ordenado": False}
        else:
            gb = _menu("AGRUPAR POR (group_by) — só o que a pergunta pede explicitamente:",
                       GROUP_BY, True, True)
            where = _montar_where()
            ordenado, order_by, limite = False, [], None
            if _sim("É RANKING? (a pergunta pede 'o maior', 'top N', 'que mais')"):
                ordenado = True
                campo = _menu("ordenar por:", [*met, *gb], False, False)[0]
                order_by = [f"-{campo}" if _sim("decrescente (do maior p/ o menor)?")
                            else campo]
                cru = input("  limite (número, ENTER p/ nenhum): ").strip()
                limite = int(cru) if cru.isdigit() else None
            spec = {"metrics": met, "group_by": gb, "where": where,
                    "order_by": order_by, "limit": limite, "ordenado": ordenado}

        print(f"  → {json.dumps(spec, ensure_ascii=False)}")
        if _sim("confirma?"):
            linhas[i]["spec_humano"] = spec
            salvar()
        else:
            print("  (descartado — o item fica pendente)")

    salvar()
    feitos = sum(1 for x in linhas if x.get("spec_humano"))
    print(f"\n{'=' * 78}\n{feitos}/{len(linhas)} anotados.")
    print("Agora: python anotar_humano.py kappa")


def kappa() -> None:
    if not FOLHA.exists():
        raise SystemExit("folha ausente — rode `python anotar_humano.py amostra` primeiro")
    linhas = [json.loads(x) for x in FOLHA.read_text(encoding="utf-8").splitlines()
              if x.strip() and not x.startswith("#")]
    preenchidos = [d for d in linhas if d.get("spec_humano")]
    if not preenchidos:
        raise SystemExit(
            "NENHUM item preenchido. O κ humano NÃO pode ser calculado por máquina — "
            "este script se recusa a inventar. Preencha 'spec_humano' à mão primeiro.")
    if len(preenchidos) < len(linhas):
        print(f"AVISO: {len(preenchidos)}/{len(linhas)} preenchidos; κ parcial.")

    autor = {it.id: it for it in _golden_selado()}
    # Guarda: um item fora do golden selado teria referência que as Fases 14/15 já descartaram.
    # Falhar alto é melhor que publicar um κ silenciosamente contaminado.
    orfaos = [d["id"] for d in preenchidos if d["id"] not in autor]
    if orfaos:
        raise SystemExit(
            f"{len(orfaos)} item(ns) anotados NÃO estão no golden selado atual: {orfaos[:5]}\n"
            "A folha é antiga (anterior às limpezas das Fases 14/15). Regenere com "
            "`python anotar_humano.py amostra` e anote a nova — o κ contra referência "
            "descartada mediria defeito conhecido, não discordância.")
    A, B = [], []
    for d in preenchidos:
        ref = autor[d["id"]]
        A.append(ref)
        spec = Spec(**d["spec_humano"])
        # O `estrato` do lado B é o julgamento DO HUMANO, não o rótulo do autor. O `ItemGolden`
        # impõe "estrato=abstencao ⟹ spec vazia" — invariante correto para um GOLDEN, mas herdar
        # `ref.estrato` aqui faria o κ ESTOURAR justamente quando o humano responde um item que o
        # autor marcou como fora-de-escopo: a discordância mais informativa da amostra (5 dos 40
        # itens são do estrato `abstencao`). Mesma convenção de `concordancia_opus5.py`;
        # `concordancia_mapeamento` só lê `.id` e `.spec`.
        estrato_b = ESTRATO_ABSTENCAO if not spec.metrics else (
            ref.estrato if ref.estrato != ESTRATO_ABSTENCAO else ESTRATOS_RESPONDIVEIS[0])
        B.append(ItemGolden(id=d["id"], pergunta_nl=ref.pergunta_nl, estrato=estrato_b,
                            spec=spec, revisado_humano=True))
    rel = concordancia_mapeamento(A, B)
    dec_a = ["fora" if not a.spec.metrics else "resp" for a in A]
    dec_b = ["fora" if not b.spec.metrics else "resp" for b in B]
    saida = carimbar({
        "tipo": "concordancia_HUMANO_x_autor_modelo",
        "n_anotados_por_humano": len(preenchidos),
        "decisao_respondivel_x_fora_kappa": cohen_kappa(dec_a, dec_b),
        "concordancia_spec": rel,
    })
    (D / "kappa_humano.json").write_text(json.dumps(saida, ensure_ascii=False, indent=2),
                                         encoding="utf-8")
    print(f"κ humano (n={len(preenchidos)}): spec bruta={rel['concordancia_spec_canonica']} "
          f"kappa_metrica={rel['cohen_kappa_metrica']}")
    print(f"-> {D / 'kappa_humano.json'}")


# ---------------------------------------------------------------------------------------------
# Planilha (XLSX) — o mesmo instrumento, para quem prefere anotar fora do terminal
#
# Vale as MESMAS garantias do anotador guiado, e cada uma custou código aqui:
#   - sem `estrato` (nem em coluna, nem no `id`: os ids são prefixados pelo estrato, então a
#     planilha carrega um `codigo` opaco e o mapa codigo->id nunca aparece para o anotador);
#   - ordem embaralhada com a MESMA semente do `anotar` (a folha é gravada agrupada por estrato);
#   - vocabulário fechado por validação de dados — o Excel só deixa escolher token que existe;
#   - nenhuma célula pré-preenchida: default seria viés silencioso.
# O `importar` revalida tudo contra as mesmas listas, porque validação de Excel é conselho, não
# garantia: dá para colar por cima dela.
# ---------------------------------------------------------------------------------------------
XLSX = G / "anotacao_humana_antt.xlsx"
COLUNAS = ["codigo", "#", "PERGUNTA", "METRICA_1", "METRICA_2", "AGRUPAR_1", "AGRUPAR_2",
           "AGRUPAR_3", "FILTRO_DIM", "FILTRO_VALOR", "RANKING", "ORDENAR_POR", "DECRESCENTE",
           "LIMITE"]
ABSTENHO = "ABSTENHO"
VAZIO = "(vazio)"


def _codigo(item_id: str) -> str:
    """Chave opaca da linha. NÃO pode ser o `id`: eles são prefixados pelo estrato
    (`abstencao_antt_23`), e uma coluna com isso entregaria a resposta na planilha."""
    return hashlib.sha1(item_id.encode()).hexdigest()[:8]


def _ordem_embaralhada(linhas: list[dict]) -> list[int]:
    ordem = list(range(len(linhas)))
    random.Random(1337).shuffle(ordem)
    return ordem


def _linhas_da_folha() -> list[dict]:
    if not FOLHA.exists():
        raise SystemExit("folha ausente — rode `python anotar_humano.py amostra` primeiro")
    return [json.loads(x) for x in FOLHA.read_text(encoding="utf-8").splitlines()
            if x.strip() and not x.startswith("#")]


def exportar_xlsx(destino: Path | None = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    destino = Path(destino) if destino else XLSX
    linhas = _linhas_da_folha()
    wb = Workbook()

    # --- listas de vocabulário (aba oculta): validação por referência de intervalo não tem o
    # limite de 255 caracteres da lista inline, e deixa o vocabulário auditável numa aba só.
    lst = wb.create_sheet("Listas")
    vocab = {
        "metrica1": [ABSTENHO, *METRICAS],
        "metrica2": [VAZIO, *METRICAS],
        "agrupar": [VAZIO, *GROUP_BY],
        "filtro_dim": [VAZIO, *sorted(VALORES)],
        "filtro_valor": [VAZIO, *[v for d in sorted(VALORES) for v in VALORES[d]]],
        "sim_nao": [VAZIO, "sim", "nao"],
        "ordenar": [VAZIO, *METRICAS, *GROUP_BY],
    }
    ref = {}
    for col, (nome, vals) in enumerate(vocab.items(), 1):
        letra = get_column_letter(col)
        lst.cell(row=1, column=col, value=nome)
        for i, v in enumerate(vals, 2):
            lst.cell(row=i, column=col, value=v)
        ref[nome] = f"Listas!${letra}$2:${letra}${len(vals) + 1}"
    lst.sheet_state = "hidden"

    # --- aba de anotação
    ws = wb.active
    ws.title = "Anotacao"
    cab = PatternFill("solid", fgColor="1F3864")
    for c, nome in enumerate(COLUNAS, 1):
        cel = ws.cell(row=1, column=c, value=nome)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = cab
    for n, i in enumerate(_ordem_embaralhada(linhas), 1):
        d = linhas[i]
        ws.cell(row=n + 1, column=1, value=_codigo(d["id"]))
        ws.cell(row=n + 1, column=2, value=n)
        cel = ws.cell(row=n + 1, column=3, value=d["pergunta_nl"])
        cel.alignment = Alignment(wrap_text=True, vertical="center")
        # NENHUMA outra célula é escrita: a planilha sai em branco de propósito.

    fim = len(linhas) + 1
    for coluna, nome in (("D", "metrica1"), ("E", "metrica2"), ("F", "agrupar"), ("G", "agrupar"),
                         ("H", "agrupar"), ("I", "filtro_dim"), ("J", "filtro_valor"),
                         ("K", "sim_nao"), ("L", "ordenar"), ("M", "sim_nao")):
        dv = DataValidation(type="list", formula1=ref[nome], allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{coluna}2:{coluna}{fim}")

    for letra, larg in (("A", 10), ("B", 5), ("C", 62), ("D", 17), ("E", 17), ("F", 22),
                        ("G", 22), ("H", 22), ("I", 22), ("J", 15), ("K", 10), ("L", 22),
                        ("M", 13), ("N", 8)):
        ws.column_dimensions[letra].width = larg
    ws.freeze_panes = "D2"

    # --- catálogo e instruções, para não precisar sair da planilha
    cat = wb.create_sheet("Catalogo")
    cat.column_dimensions["A"].width = 100
    for i, ln in enumerate(CATALOGO_ANTT.splitlines(), 1):
        cat.cell(row=i, column=1, value=ln)

    ins = wb.create_sheet("Instrucoes")
    ins.column_dimensions["A"].width = 100
    texto = [
        "COMO PREENCHER",
        "",
        "Uma linha por pergunta. Todas as colunas coloridas tem lista suspensa — clique e escolha.",
        "",
        "METRICA_1   qual metrica do catalogo responde. Se NENHUMA responde, escolha ABSTENHO",
        "            e deixe o resto da linha em branco: o item acabou.",
        "METRICA_2   so se a pergunta pedir DUAS metricas. Normalmente fica vazio.",
        "AGRUPAR_1/2/3   'por X' = agrupar. So o que a pergunta pede EXPLICITAMENTE.",
        "FILTRO_DIM + FILTRO_VALOR   'dos X' = filtrar por um valor especifico.",
        "            Filtrar NAO e agrupar: 'veiculos comerciais' vai em FILTRO, nao em AGRUPAR.",
        "RANKING     'sim' so se a pergunta pedir ordem explicita: 'o maior', 'top 5', 'que mais'.",
        "            'por mes' e 'por praca' NAO sao ranking.",
        "ORDENAR_POR / DECRESCENTE / LIMITE   so preencha se RANKING = sim.",
        "",
        "REGRA PRATICA:  'por X' agrupa  |  'dos X' filtra.",
        "",
        "NAO MEXA na coluna 'codigo' — e ela que devolve cada resposta ao item certo.",
        "Nao acrescente nem apague linhas.",
        "",
        "EM DUVIDA, anote sua melhor leitura e siga. Discordar do outro anotador e o SINAL que",
        "estamos medindo — nao e erro seu, e nao deve ser evitado. Nao tente adivinhar o que o",
        "outro respondeu: uma concordancia obtida assim nao vale nada.",
        "",
        "NAO ABRA golden/golden_test_antt.jsonl — e onde estao as respostas do outro anotador.",
        "",
        "Pode parar no meio e continuar depois. O que ficar em branco continua pendente.",
        "",
        "AO TERMINAR, salve e rode:",
        "    .venv/bin/python anotar_humano.py importar <caminho-da-planilha>",
        "    .venv/bin/python anotar_humano.py kappa",
    ]
    for i, ln in enumerate(texto, 1):
        cel = ins.cell(row=i, column=1, value=ln)
        if i == 1 or ln.startswith(("REGRA", "NAO ABRA", "AO TERMINAR")):
            cel.font = Font(bold=True)

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    print(f"planilha gerada: {len(linhas)} itens (ordem embaralhada, sem estrato) -> {destino}")
    print("Preencha e depois rode: python anotar_humano.py importar <caminho>")


def importar_xlsx(origem: Path) -> None:
    """Lê a planilha preenchida de volta para a folha. Revalida TUDO: a validação do Excel é
    conselho (dá para colar por cima), e um token invalido viraria ruído dentro do κ."""
    from openpyxl import load_workbook

    origem = Path(origem)
    if not origem.exists():
        raise SystemExit(f"planilha não encontrada: {origem}")
    linhas = _linhas_da_folha()
    por_codigo = {_codigo(d["id"]): i for i, d in enumerate(linhas)}

    ws = load_workbook(origem, data_only=True)["Anotacao"]
    cabecalho = [c.value for c in ws[1]]
    if cabecalho[:len(COLUNAS)] != COLUNAS:
        raise SystemExit(f"cabeçalho inesperado — a planilha foi alterada?\n  {cabecalho}")

    def txt(v: object) -> str:
        return str(v).strip() if v is not None and str(v).strip() not in ("", VAZIO) else ""

    novos, erros = 0, []
    for n, linha in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        reg = dict(zip(COLUNAS, linha[:len(COLUNAS)], strict=False))
        cod = txt(reg["codigo"])
        if not cod:
            continue
        if cod not in por_codigo:
            erros.append(f"linha {n}: codigo '{cod}' não corresponde a nenhum item da folha")
            continue
        m1 = txt(reg["METRICA_1"])
        if not m1:
            continue                      # em branco = ainda pendente, não é erro
        if m1 == ABSTENHO:
            spec = {"metrics": [], "group_by": [], "where": None,
                    "order_by": [], "limit": None, "ordenado": False}
        else:
            met = [m for m in (m1, txt(reg["METRICA_2"])) if m]
            gb = [g for g in (txt(reg["AGRUPAR_1"]), txt(reg["AGRUPAR_2"]),
                              txt(reg["AGRUPAR_3"])) if g]
            if any(m not in METRICAS for m in met):
                erros.append(f"linha {n}: métrica fora do catálogo: {met}")
                continue
            if any(g not in GROUP_BY for g in gb):
                erros.append(f"linha {n}: token de agrupamento fora do catálogo: {gb}")
                continue
            dim, val = txt(reg["FILTRO_DIM"]), txt(reg["FILTRO_VALOR"])
            if bool(dim) != bool(val):
                erros.append(f"linha {n}: FILTRO_DIM e FILTRO_VALOR têm de vir juntos")
                continue
            if dim and val not in VALORES.get(dim, []):
                erros.append(f"linha {n}: '{val}' não é valor válido de {dim}")
                continue
            where = f"{{{{ Dimension('{dim}') }}}} = '{val}'" if dim else None
            ordenado = txt(reg["RANKING"]).lower() == "sim"
            order_by, limite = [], None
            if ordenado:
                campo = txt(reg["ORDENAR_POR"])
                if campo not in [*met, *gb]:
                    erros.append(f"linha {n}: ORDENAR_POR '{campo}' não está entre as métricas "
                                 "nem os agrupamentos desta linha")
                    continue
                desc = txt(reg["DECRESCENTE"]).lower() != "nao"
                order_by = [f"-{campo}" if desc else campo]
                bruto = txt(reg["LIMITE"])
                limite = int(float(bruto)) if bruto.replace(".", "").isdigit() else None
            spec = {"metrics": met, "group_by": gb, "where": where,
                    "order_by": order_by, "limit": limite, "ordenado": ordenado}
        linhas[por_codigo[cod]]["spec_humano"] = spec
        novos += 1

    if erros:
        raise SystemExit("planilha REJEITADA — nada foi gravado:\n  " + "\n  ".join(erros))

    cab = [x for x in FOLHA.read_text(encoding="utf-8").splitlines() if x.startswith("#")]
    FOLHA.write_text("\n".join(cab) + "\n"
                     + "\n".join(json.dumps(d, ensure_ascii=False) for d in linhas) + "\n",
                     encoding="utf-8")
    print(f"importados {novos} itens; folha agora com "
          f"{sum(1 for d in linhas if d.get('spec_humano'))}/{len(linhas)} anotados.")
    print("Agora: python anotar_humano.py kappa")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "amostra"
    if cmd == "amostra":
        amostra(int(sys.argv[2]) if len(sys.argv) > 2 else 40)
    elif cmd == "anotar":
        anotar()
    elif cmd == "xlsx":
        exportar_xlsx(Path(sys.argv[2]) if len(sys.argv) > 2 else None)
    elif cmd == "importar":
        if len(sys.argv) < 3:
            raise SystemExit("uso: python anotar_humano.py importar <caminho-da-planilha>")
        importar_xlsx(Path(sys.argv[2]))
    elif cmd == "kappa":
        kappa()
    else:
        print(__doc__)
